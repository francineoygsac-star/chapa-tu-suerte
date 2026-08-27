
import os, secrets, hmac, time, mimetypes, json, urllib.request, urllib.error, re, unicodedata, hashlib, base64
from collections import defaultdict, deque
from io import BytesIO
from datetime import datetime, timezone
from zoneinfo import ZoneInfo
from pathlib import Path
from functools import wraps
from flask import Flask, request, jsonify, send_from_directory, make_response, redirect, session, send_file, g
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from google.cloud import storage
from google.cloud import firestore
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.security import generate_password_hash, check_password_hash
from cryptography.fernet import Fernet, InvalidToken

BASE = Path(__file__).resolve().parent
PUBLIC = BASE / "public"

app = Flask(__name__, static_folder=str(PUBLIC), static_url_path="")
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)
app.config["MAX_CONTENT_LENGTH"]=25*1024*1024
app.secret_key = os.environ.get("FLASK_SECRET", "CHANGE_THIS_SECRET_IN_PRODUCTION")
ADMIN_USER = os.environ.get("ADMIN_USER", "admin")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "CAMBIA_ESTA_CLAVE")
ADMIN_TOTP_SECRET = os.environ.get("ADMIN_TOTP_SECRET", "").replace(" ", "").strip().upper()
MAX_TICKETS = 999999
PRICE = 10
JSONPE_TOKEN = os.environ.get("JSONPE_TOKEN", "").strip()
GCS_BUCKET_NAME = os.environ.get("GCS_BUCKET_NAME", "").strip()
DNI_CACHE = {}
DNI_CACHE_SECONDS = 90*24*60*60
_FIRESTORE_DB = None
_ADMIN_SESSION_CACHE = {}
_LAST_EXPIRED_CLEANUP = 0
limiter = Limiter(key_func=get_remote_address, app=app, default_limits=[], storage_uri="memory://")

@app.errorhandler(429)
def rate_limited(_error):
    return jsonify(error="Demasiados intentos. Espera unos minutos antes de volver a intentar."),429

def token_digest(value):
    return hashlib.sha256(str(value).encode("utf-8")).hexdigest()

def credential_cipher():
    """Cifra los secretos TOTP de usuarios secundarios antes de guardarlos."""
    key=base64.urlsafe_b64encode(hashlib.sha256(app.secret_key.encode("utf-8")).digest())
    return Fernet(key)

def encrypt_totp(secret):
    return credential_cipher().encrypt(secret.encode("ascii")).decode("ascii")

def decrypt_totp(encrypted):
    return credential_cipher().decrypt(str(encrypted).encode("ascii")).decode("ascii")

def audit(action, details=None, actor=None):
    """Registra acciones sensibles sin guardar contraseñas, tokens ni comprobantes."""
    try:
        firestore_db().collection("admin_audit").add({
            "action":str(action)[:80], "actor":str(actor or getattr(g,"admin_user",None) or ADMIN_USER)[:80],
            "ip":get_remote_address(), "created_at":int(time.time()),
            "details":details or {},
        })
    except Exception:
        app.logger.exception("No se pudo guardar el evento de auditoría")

def same_origin_request():
    origin=(request.headers.get("Origin") or "").rstrip("/")
    referer=(request.headers.get("Referer") or "").rstrip("/")
    expected=request.host_url.rstrip("/")
    return (not origin or hmac.compare_digest(origin,expected)) and (not referer or referer.startswith(expected+"/"))

@app.after_request
def security_headers(response):
    response.headers.setdefault("X-Content-Type-Options","nosniff")
    response.headers.setdefault("X-Frame-Options","DENY")
    response.headers.setdefault("Referrer-Policy","no-referrer")
    response.headers.setdefault("Permissions-Policy","camera=(), microphone=(), geolocation=()")
    response.headers.setdefault("Content-Security-Policy",
        "default-src 'self'; img-src 'self' data:; style-src 'self' 'unsafe-inline'; "
        "script-src 'self' 'unsafe-inline'; connect-src 'self'; object-src 'none'; "
        "base-uri 'self'; frame-ancestors 'none'; form-action 'self'")
    if request.is_secure:
        response.headers.setdefault("Strict-Transport-Security","max-age=31536000; includeSubDomains")
    if request.path.startswith("/admin") or request.path.startswith("/api/admin"):
        response.headers["Cache-Control"]="no-store, private"
    return response

def firestore_db():
    global _FIRESTORE_DB
    if _FIRESTORE_DB is None:
        _FIRESTORE_DB = firestore.Client(database="(default)")
    return _FIRESTORE_DB

def storage_bucket():
    if not GCS_BUCKET_NAME:
        raise RuntimeError("GCS_BUCKET_NAME no está configurado")
    return storage.Client().bucket(GCS_BUCKET_NAME)

def safe_filename_part(value):
    value=unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    value=re.sub(r"[^A-Za-z0-9]+", "_", value).strip("_").upper()
    return value[:60] or "SIN_NOMBRE"

def lima_datetime_text(timestamp):
    """Convierte un timestamp UTC a la fecha y hora administrativa de Perú."""
    if not timestamp:
        return ""
    return datetime.fromtimestamp(int(timestamp),timezone.utc).astimezone(
        ZoneInfo("America/Lima")
    ).strftime("%Y-%m-%d %H:%M:%S")

def cleanup_expired():
    """Libera reservas pendientes con más de 30 minutos."""
    global _LAST_EXPIRED_CLEANUP
    now=int(time.time())
    # Evita repetir la misma consulta de mantenimiento en cada carga del panel.
    if now-_LAST_EXPIRED_CLEANUP < 60:
        return
    _LAST_EXPIRED_CLEANUP=now
    cutoff=now-1800
    pending=firestore_db().collection("orders").where("status", "==", "pending").limit(200).stream()
    for snapshot in pending:
        order=snapshot.to_dict()
        if int(order.get("created_at", now)) >= cutoff:
            continue
        batch=firestore_db().batch()
        batch.update(snapshot.reference, {
            "status":"expired", "reviewed_at":now,
            "reviewed_at_text":time.strftime("%Y-%m-%d %H:%M:%S", time.gmtime(now)),
            "review_note":"Reserva vencida", "released_tickets":order.get("tickets",[]), "tickets":[],
        })
        for number in order.get("tickets", []):
            batch.delete(firestore_db().collection("tickets").document(str(number)))
        batch.commit()

def make_code():
    return "CTS-" + secrets.token_hex(4).upper()

def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        token=request.cookies.get("admin_session")
        if not token: return jsonify(error="No autorizado"),401
        now=int(time.time())
        digest=token_digest(token)
        cached=_ADMIN_SESSION_CACHE.get(digest) or {}
        if int(cached.get("cached_until",0)) <= now:
            snapshot=firestore_db().collection("admin_sessions").document(digest).get()
            if not snapshot.exists or int((snapshot.to_dict() or {}).get("created_at",0)) <= now-86400:
                _ADMIN_SESSION_CACHE.pop(digest,None)
                return jsonify(error="No autorizado"),401
            saved=snapshot.to_dict() or {}
            cached={"cached_until":now+300,"csrf_hash":saved.get("csrf_hash",""),
                    "actor":saved.get("actor",ADMIN_USER),"role":saved.get("role","owner")}
            _ADMIN_SESSION_CACHE[digest]=cached
        g.admin_user=str(cached.get("actor") or ADMIN_USER)
        g.admin_role=str(cached.get("role") or "viewer")
        if request.method not in {"GET","HEAD","OPTIONS"}:
            csrf=request.headers.get("X-CSRF-Token","")
            if (not same_origin_request() or not csrf or
                    not hmac.compare_digest(token_digest(csrf),str(cached.get("csrf_hash","")))):
                return jsonify(error="Solicitud de seguridad inválida. Vuelve a iniciar sesión."),403
        return fn(*args, **kwargs)
    return wrapper

def roles_allowed(*roles):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args,**kwargs):
            if getattr(g,"admin_role",None) not in roles:
                return jsonify(error="No tienes permiso para realizar esta acción"),403
            return fn(*args,**kwargs)
        return wrapper
    return decorator

@app.get("/")
def home():
    return send_from_directory(PUBLIC,"index.html")

@app.post("/api/identity/dni")
@limiter.limit("12 per minute; 40 per hour")
def identity_dni():
    """Consulta DNI en JSON.pe sin exponer el token al navegador."""
    dni=(request.json or {}).get("dni", "") if request.is_json else (request.form.get("dni") or "")
    dni=str(dni).strip()
    if not dni.isdigit() or len(dni)!=8:
        return jsonify(error="El DNI debe tener 8 dígitos"),400
    if not JSONPE_TOKEN:
        return jsonify(error="La consulta de DNI no está configurada en el servidor"),503

    cached=DNI_CACHE.get(dni)
    if cached:
        return jsonify(success=True, data=cached, cached=True, cache_source="memory")

    # Caché permanente: sobrevive a suspensiones, reinicios y despliegues de Render.
    now=int(time.time())
    cache_ref=firestore_db().collection("dni_cache").document(dni)
    try:
        cache_snapshot=cache_ref.get()
        if cache_snapshot.exists:
            saved=cache_snapshot.to_dict() or {}
            cache_age=now-int(saved.get("updated_at",0))
            if saved.get("status")=="not_found" and cache_age < 24*60*60:
                return jsonify(error="No se encontraron datos para este DNI. Escribe el nombre completo manualmente.",cached=True),404
            if saved.get("status")=="temporary_error" and cache_age < 5*60:
                return jsonify(error="El servicio de consulta no respondió correctamente. Escribe el nombre completo manualmente o intenta más tarde.",cached=True),503
            if (int(saved.get("updated_at",0)) >= now-DNI_CACHE_SECONDS
                    and str(saved.get("nombre_completo") or "").strip()):
                clean={key:saved.get(key) for key in (
                    "numero","nombres","apellido_paterno","apellido_materno",
                    "nombre_completo","codigo_verificacion"
                )}
                clean["numero"]=dni
                DNI_CACHE[dni]=clean
                return jsonify(success=True,data=clean,cached=True,cache_source="firestore")
    except Exception:
        # Si el caché no está disponible, la validación oficial todavía puede continuar.
        pass

    payload=json.dumps({"dni":dni}).encode("utf-8")
    req=urllib.request.Request(
        "https://api.json.pe/api/dni",
        data=payload,
        headers={
            "Authorization": f"Bearer {JSONPE_TOKEN}",
            "Content-Type": "application/json",
            "Accept": "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            raw=resp.read().decode("utf-8")
            result=json.loads(raw)
    except urllib.error.HTTPError as e:
        try:
            detail=json.loads(e.read().decode("utf-8"))
            message=detail.get("message") or detail.get("error") or "No se pudo consultar el DNI"
        except Exception:
            message="No se pudo consultar el DNI"
        if e.code in {404,422}:
            try: cache_ref.set({"numero":dni,"status":"not_found","updated_at":now})
            except Exception: pass
            message="No se encontraron datos para este DNI. Escribe el nombre completo manualmente."
        return jsonify(error=message), e.code if 400 <= e.code < 500 else 502
    except (urllib.error.URLError, TimeoutError, json.JSONDecodeError):
        # Evita que un doble clic o una respuesta defectuosa consuma de nuevo
        # durante los siguientes cinco minutos.
        try: cache_ref.set({"numero":dni,"status":"temporary_error","updated_at":now})
        except Exception: pass
        return jsonify(error="El servicio de consulta no respondió correctamente. Escribe el nombre completo manualmente o intenta más tarde."),502

    if not result.get("success"):
        try: cache_ref.set({"numero":dni,"status":"not_found","updated_at":now})
        except Exception: pass
        return jsonify(error="No se encontraron datos para este DNI. Escribe el nombre completo manualmente."),404

    data=result.get("data") or {}
    nombres=str(data.get("nombres") or "").strip()
    paterno=str(data.get("apellido_paterno") or "").strip()
    materno=str(data.get("apellido_materno") or "").strip()
    nombre_completo=str(data.get("nombre_completo") or " ".join(x for x in [nombres,paterno,materno] if x)).strip()
    if not nombre_completo:
        return jsonify(error="La consulta no devolvió nombres y apellidos"),404

    clean={
        "numero": dni,
        "nombres": nombres,
        "apellido_paterno": paterno,
        "apellido_materno": materno,
        "nombre_completo": nombre_completo,
        "codigo_verificacion": data.get("codigo_verificacion"),
    }
    DNI_CACHE[dni]=clean
    try:
        cache_ref.set({**clean,"status":"verified","updated_at":now})
    except Exception:
        # No se invalida una consulta oficial exitosa por un fallo al guardar el caché.
        pass
    return jsonify(success=True, data=clean, cached=False, cache_source="jsonpe")


@app.post("/api/orders")
@limiter.limit("5 per 10 minutes; 20 per day")
def create_order():
    cleanup_expired()
    name=(request.form.get("name") or "").strip()
    phone=re.sub(r"\D", "", request.form.get("phone") or "")
    document_type=(request.form.get("document_type") or "").strip().upper()
    document_number=(request.form.get("document_number") or "").strip()
    try: quantity=int(request.form.get("quantity","1"))
    except: quantity=0
    proofs=[item for item in request.files.getlist("proof") if item and item.filename]
    if not name or len(name)>120: return jsonify(error="Nombre inválido"),400
    if not re.fullmatch(r"9\d{8}", phone):
        return jsonify(error="Número de WhatsApp inválido"),400
    if document_type not in {"DNI","CE"}: return jsonify(error="Tipo de documento inválido"),400
    if not document_number.isdigit(): return jsonify(error="Número de documento inválido"),400
    if document_type=="DNI" and len(document_number)!=8: return jsonify(error="El DNI debe tener 8 dígitos"),400
    if document_type=="CE" and not 6<=len(document_number)<=12: return jsonify(error="El Carnet de Extranjería debe tener entre 6 y 12 dígitos"),400
    if quantity<1 or quantity>50: return jsonify(error="Cantidad inválida"),400
    if not proofs: return jsonify(error="Falta el comprobante"),400
    if len(proofs)>5: return jsonify(error="Puedes adjuntar como máximo 5 comprobantes"),400
    normalized_proofs=[]
    for proof in proofs:
        if proof.mimetype not in {"image/jpeg","image/png","image/webp"}:
            return jsonify(error="Todos los comprobantes deben ser imágenes JPG, PNG o WEBP"),400
        raw=proof.read()
        if len(raw)>5*1024*1024: return jsonify(error="Cada comprobante puede pesar como máximo 5 MB"),400
        try:
            from PIL import Image
            Image.MAX_IMAGE_PIXELS=25_000_000
            source=Image.open(BytesIO(raw)); source.verify(); source=Image.open(BytesIO(raw))
            if getattr(source,"is_animated",False):
                return jsonify(error="No se permiten imágenes animadas"),400
            source.load(); output=BytesIO()
            if proof.mimetype=="image/png":
                source.convert("RGBA" if "A" in source.getbands() else "RGB").save(output,"PNG",optimize=True)
                ext=".png"; verified_type="image/png"
            elif proof.mimetype=="image/webp":
                source.convert("RGB").save(output,"WEBP",quality=88,method=4)
                ext=".webp"; verified_type="image/webp"
            else:
                source.convert("RGB").save(output,"JPEG",quality=90,optimize=True)
                ext=".jpg"; verified_type="image/jpeg"
            normalized_proofs.append((output.getvalue(),ext,verified_type))
        except Exception:
            return jsonify(error="Uno de los archivos no es una imagen válida"),400

    code=make_code()
    oid=str(int(time.time()*1000)*100+secrets.randbelow(100))
    now=int(time.time())
    created_at_text=time.strftime("%Y-%m-%d %H:%M:%S", time.gmtime(now))
    orders=firestore_db().collection("orders")
    tickets=firestore_db().collection("tickets")
    nums=[]
    uploaded_proofs=[]
    try:
        # Reserva todos los números y crea la solicitud en una sola transacción.
        for _ in range(12):
            candidates=[]
            while len(candidates)<quantity:
                n=secrets.randbelow(900000)+100000
                if n not in candidates: candidates.append(n)

            transaction=firestore_db().transaction()
            @firestore.transactional
            def reserve(transaction):
                refs=[tickets.document(str(n)) for n in candidates]
                snapshots=[ref.get(transaction=transaction) for ref in refs]
                if any(snapshot.exists for snapshot in snapshots):
                    return False
                transaction.set(orders.document(oid), {
                    "id":oid, "code":code, "name":name, "phone":phone,
                    "document_type":document_type, "document_number":document_number,
                    "quantity":quantity, "total":quantity*PRICE,
                    "proof_path":None,"proof_paths":[], "status":"pending", "tickets":candidates,
                    "created_at":now, "created_at_text":created_at_text,
                    "reviewed_at":None, "reviewed_at_text":None, "review_note":None,
                })
                for ref,n in zip(refs,candidates):
                    transaction.set(ref, {
                        "number":n, "order_id":oid, "order_code":code,
                        "status":"reserved", "created_at":now,
                    })
                return True
            if reserve(transaction):
                nums=candidates
                break
        if not nums:
            raise RuntimeError("No se pudieron reservar números únicos")

        # Los nombres no contienen DNI, teléfono ni nombre del comprador.
        for raw,ext,verified_type in normalized_proofs:
            filename=f"proof_{secrets.token_hex(24)}{ext}"
            blob=storage_bucket().blob(filename)
            blob.upload_from_string(raw, content_type=verified_type)
            uploaded_proofs.append(filename)
        orders.document(oid).update({"proof_path":uploaded_proofs[0],
                                     "proof_paths":uploaded_proofs})
    except Exception:
        # Si falla la imagen, libera la reserva para no dejar tickets bloqueados.
        if nums:
            batch=firestore_db().batch()
            batch.delete(orders.document(oid))
            for number in nums:
                batch.delete(tickets.document(str(number)))
            batch.commit()
        for filename in uploaded_proofs:
            try: storage_bucket().blob(filename).delete()
            except Exception: pass
        return jsonify(error="No se pudo crear la solicitud"),500
    return jsonify(order_code=code,quantity=quantity,total=quantity*PRICE,tickets=nums,
                   proof_count=len(uploaded_proofs),status="pending"),201

@app.get("/api/tickets/<number>")
@limiter.limit("30 per minute")
def ticket(number):
    if not number.isdigit() or len(number)!=6: return jsonify(error="Número inválido"),400
    ticket_snapshot=firestore_db().collection("tickets").document(number).get()
    if not ticket_snapshot.exists: return jsonify(error="No encontrado"),404
    ticket_data=ticket_snapshot.to_dict()
    order_snapshot=firestore_db().collection("orders").document(ticket_data["order_id"]).get()
    if not order_snapshot.exists: return jsonify(error="No encontrado"),404
    order=order_snapshot.to_dict()
    status_map={
      ("reserved","pending"):("RESERVADO · PENDIENTE DE PAGO/VALIDACIÓN"),
      ("reserved","expired"):("RESERVA VENCIDA"),
      ("reserved","rejected"):("RECHAZADO"),
      ("confirmed","approved"):("CONFIRMADO"),
    }
    label=status_map.get((ticket_data["status"],order["status"]),"EN REVISIÓN")
    words=str(order.get("name") or "").split()
    masked_name=" ".join((word[:1]+"***") for word in words[:3])
    return jsonify(number=ticket_data["number"],name=masked_name,status_label=label)


@app.get("/api/tickets/by-document")
@limiter.limit("10 per minute; 50 per day")
def tickets_by_document():
    document_type=(request.args.get("document_type") or "").strip().upper()
    document_number=(request.args.get("document_number") or "").strip()
    phone=re.sub(r"\D","",request.args.get("phone") or "")
    if document_type not in {"DNI","CE"} or not document_number.isdigit():
        return jsonify(error="Documento inválido"),400
    if document_type=="DNI" and len(document_number)!=8:
        return jsonify(error="El DNI debe tener 8 dígitos"),400
    if document_type=="CE" and not 6<=len(document_number)<=12:
        return jsonify(error="El Carnet de Extranjería debe tener entre 6 y 12 dígitos"),400
    if not re.fullmatch(r"9\d{8}",phone):
        return jsonify(error="Ingresa el WhatsApp usado en la compra"),400
    cleanup_expired()
    snapshots=firestore_db().collection("orders").where("document_number", "==", document_number).limit(200).stream()
    orders=[snapshot.to_dict() for snapshot in snapshots]
    orders=[o for o in orders if o.get("document_type")==document_type and
            hmac.compare_digest(str(o.get("phone") or ""),phone)]
    if not orders: return jsonify(error="No se encontraron tickets para ese documento"),404
    labels={"pending":"PENDIENTE DE VALIDACIÓN","approved":"CONFIRMADO","rejected":"RECHAZADO","expired":"RESERVA VENCIDA"}
    orders.sort(key=lambda o:int(o.get("created_at",0)), reverse=True)
    result=[]
    for order in orders:
        ticket_status="confirmed" if order.get("status")=="approved" else "reserved"
        result.append({
            "order_code":order["code"], "name":order["name"],
            "document_type":order["document_type"], "document_number":order["document_number"],
            "quantity":order["quantity"], "total":order["total"],
            "status":order["status"], "status_label":labels.get(order["status"],"EN REVISIÓN"),
            "created_at":order.get("created_at_text",""),
            "tickets":[{"number":n,"status":ticket_status} for n in sorted(order.get("tickets",[]))]
        })
    return jsonify(document_type=document_type,document_number=document_number,total_orders=len(result),
                   total_tickets=sum(len(x["tickets"]) for x in result),orders=result)

@app.post("/admin/login")
@limiter.limit("5 per minute; 20 per hour")
def login():
    data=request.form or request.json or {}
    if ADMIN_PASSWORD=="CAMBIA_ESTA_CLAVE" or app.secret_key=="CHANGE_THIS_SECRET_IN_PRODUCTION":
        return jsonify(error="La seguridad del administrador no está configurada"),503
    import pyotp
    supplied_user=str(data.get("username","")).strip()
    supplied_password=str(data.get("password",""))
    supplied_totp=str(data.get("totp","")).strip()
    actor=None; role=None; valid_password=False; valid_totp=False; user_ref=None

    normalized=supplied_user.lower()
    if re.fullmatch(r"[a-z0-9._-]{3,32}",normalized):
        user_ref=firestore_db().collection("admin_users").document(normalized)
        user_snapshot=user_ref.get()
        if user_snapshot.exists:
            user_data=user_snapshot.to_dict() or {}
            actor=str(user_data.get("username") or normalized)
            role=str(user_data.get("role") or "viewer")
            valid_password=bool(user_data.get("active",False)) and check_password_hash(
                str(user_data.get("password_hash") or ""),supplied_password)
            try:
                secret=decrypt_totp(user_data.get("totp_secret_encrypted",""))
                valid_totp=pyotp.TOTP(secret).verify(supplied_totp,valid_window=1)
            except Exception:
                valid_totp=False

    # La cuenta original de Render permanece como administrador principal.
    if actor is None and hmac.compare_digest(supplied_user,ADMIN_USER):
        actor=ADMIN_USER; role="owner"; user_ref=None
        valid_password=hmac.compare_digest(supplied_password,ADMIN_PASSWORD)
        try:
            valid_totp=bool(ADMIN_TOTP_SECRET) and pyotp.TOTP(ADMIN_TOTP_SECRET).verify(
                supplied_totp,valid_window=1)
        except Exception:
            valid_totp=False
    if not (actor and valid_password and valid_totp):
        audit("login_failed",{"username":str(data.get("username", ""))[:80]},actor="unknown")
        return jsonify(error="Credenciales incorrectas"),401
    if role not in {"owner","validator","viewer"}:
        return jsonify(error="Credenciales incorrectas"),401
    token=secrets.token_urlsafe(32)
    csrf=secrets.token_urlsafe(32)
    digest=token_digest(token)
    now=int(time.time())
    saved={"created_at":now,"csrf_hash":token_digest(csrf),"actor":actor,"role":role}
    firestore_db().collection("admin_sessions").document(digest).set(saved)
    _ADMIN_SESSION_CACHE[digest]={"cached_until":now+300,"csrf_hash":saved["csrf_hash"],
                                  "actor":actor,"role":role}
    if user_ref is not None:
        user_ref.update({"last_login_at":now})
    audit("login_success",{"role":role},actor=actor)
    r=jsonify(ok=True,csrf_token=csrf,totp_enabled=True,user=actor,role=role)
    r.set_cookie("admin_session",token,httponly=True,samesite="Strict",secure=True,max_age=86400,path="/")
    return r

@app.post("/admin/logout")
@admin_required
def logout():
    token=request.cookies.get("admin_session")
    if token:
        digest=token_digest(token)
        _ADMIN_SESSION_CACHE.pop(digest,None)
        firestore_db().collection("admin_sessions").document(digest).delete()
    audit("logout")
    r=jsonify(ok=True); r.delete_cookie("admin_session",path="/"); return r

@app.post("/api/admin/sessions/revoke-all")
@admin_required
@roles_allowed("owner")
def revoke_all_sessions():
    sessions=list(firestore_db().collection("admin_sessions").limit(500).stream())
    batch=firestore_db().batch()
    for snapshot in sessions:
        batch.delete(snapshot.reference)
    batch.commit()
    _ADMIN_SESSION_CACHE.clear()
    audit("all_sessions_revoked",{"count":len(sessions)})
    response=jsonify(ok=True,count=len(sessions))
    response.delete_cookie("admin_session",path="/")
    return response

@app.get("/api/admin/orders")
@admin_required
def admin_orders():
    cleanup_expired()
    status=request.args.get("status","all")
    query=firestore_db().collection("orders")
    if status!="all": query=query.where("status", "==", status)
    rows=[snapshot.to_dict() for snapshot in query.limit(200).stream()]
    rows.sort(key=lambda row:int(row.get("created_at",0)), reverse=True)
    for row in rows:
        row["created_at"]=lima_datetime_text(row.get("created_at")) or row.get("created_at_text","")
        row["tickets"]=sorted(row.get("tickets",[]) or row.get("released_tickets",[]))
        proof_paths=row.get("proof_paths") or ([row.get("proof_path")] if row.get("proof_path") else [])
        row["proof_urls"]=["/admin/proof/"+path for path in proof_paths if path]
        row["proof_url"]=row["proof_urls"][0] if row["proof_urls"] else None
    return jsonify(orders=rows)

@app.get("/api/admin/export.xlsx")
@admin_required
@roles_allowed("owner","validator")
def export_admin_excel():
    # Estas librerías son pesadas; se cargan solo cuando se descarga el Excel,
    # no durante el arranque normal de la web ni del panel.
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    rows=[snapshot.to_dict() for snapshot in firestore_db().collection("orders").stream()]
    audit("export_excel",{"records":len(rows)})
    rows.sort(key=lambda row:int(row.get("created_at",0)), reverse=True)
    status_labels={"pending":"Pendiente","approved":"Aprobada","rejected":"Rechazada","expired":"Vencida"}
    lima=ZoneInfo("America/Lima")

    def local_datetime(timestamp):
        if not timestamp: return None
        return datetime.fromtimestamp(int(timestamp),timezone.utc).astimezone(lima).replace(tzinfo=None)

    def safe_excel_text(value):
        text=str(value or "")
        return "'"+text if text.startswith(("=","+","-","@")) else text

    workbook=Workbook()
    summary=workbook.active
    summary.title="Resumen"
    detail=workbook.create_sheet("Participantes y pagos")
    yellow="FFD400"; dark="111117"; white="FFFFFF"; light="F2F2F2"
    thin=Side(style="thin",color="D9D9D9")
    summary.sheet_view.showGridLines=False
    detail.sheet_view.showGridLines=False

    approved=[row for row in rows if row.get("status")=="approved"]
    pending=[row for row in rows if row.get("status")=="pending"]
    rejected=[row for row in rows if row.get("status")=="rejected"]
    expired=[row for row in rows if row.get("status")=="expired"]
    confirmed_tickets=sum(int(row.get("quantity",0)) for row in approved)
    approved_sales=sum(int(row.get("total",0)) for row in approved)
    reactivated=sum(bool(row.get("reactivated_from_expired")) for row in rows)

    summary.merge_cells("A1:H2")
    summary["A1"]="CHAPA TU SUERTE — REPORTE ADMINISTRATIVO"
    summary["A1"].font=Font(name="Calibri",bold=True,size=20,color=white)
    summary["A1"].fill=PatternFill("solid",fgColor=dark)
    summary["A1"].alignment=Alignment(horizontal="center",vertical="center")
    summary.merge_cells("A3:H3")
    summary["A3"]="Generado el "+datetime.now(lima).strftime("%d/%m/%Y %H:%M")+" · Hora de Perú"
    summary["A3"].font=Font(italic=True,color="D1D5DB")
    summary["A3"].fill=PatternFill("solid",fgColor="24242D")
    summary["A3"].alignment=Alignment(horizontal="center")

    top_labels=["TOTAL SOLICITUDES","PENDIENTES","APROBADAS","RECHAZADAS"]
    top_values=[len(rows),len(pending),len(approved),len(rejected)]
    bottom_labels=["VENTAS APROBADAS","TICKETS CONFIRMADOS","VENCIDAS","REACTIVADAS"]
    bottom_values=[approved_sales,confirmed_tickets,len(expired),reactivated]
    for column,(label,value) in enumerate(zip(top_labels,top_values),1):
        summary.cell(5,column,label); summary.cell(6,column,value)
    for column,(label,value) in enumerate(zip(bottom_labels,bottom_values),1):
        summary.cell(8,column,label); summary.cell(9,column,value)
    for row_number in (5,8):
        for cell in summary[row_number][:4]:
            cell.font=Font(bold=True,color=white,size=11)
            cell.fill=PatternFill("solid",fgColor="24242D")
            cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    for row_number in (6,9):
        for cell in summary[row_number][:4]:
            cell.font=Font(bold=True,color=dark,size=22)
            cell.fill=PatternFill("solid",fgColor=light)
            cell.alignment=Alignment(horizontal="center",vertical="center")
            cell.border=Border(bottom=thin)
    summary["A9"].number_format='"S/" #,##0'
    summary.row_dimensions[5].height=26; summary.row_dimensions[6].height=46
    summary.row_dimensions[8].height=26; summary.row_dimensions[9].height=46

    summary.append([])
    status_summary=[
        ["ESTADO","SOLICITUDES","TICKETS","VENTAS (S/)"],
        ["Aprobada",len(approved),confirmed_tickets,approved_sales],
        ["Pendiente",len(pending),sum(int(row.get("quantity",0)) for row in pending),sum(int(row.get("total",0)) for row in pending)],
        ["Rechazada",len(rejected),sum(int(row.get("quantity",0)) for row in rejected),sum(int(row.get("total",0)) for row in rejected)],
        ["Vencida",len(expired),sum(int(row.get("quantity",0)) for row in expired),sum(int(row.get("total",0)) for row in expired)],
    ]
    for row_index,row_values in enumerate(status_summary,11):
        for column,value in enumerate(row_values,1): summary.cell(row_index,column,value)
    for cell in summary[11][:4]:
        cell.font=Font(bold=True,color=dark); cell.fill=PatternFill("solid",fgColor=yellow)
        cell.alignment=Alignment(horizontal="center")
    for row_number in range(12,16):
        summary.cell(row_number,1).font=Font(bold=True,color="1F2937")
        summary.cell(row_number,4).number_format='"S/" #,##0'
        for cell in summary[row_number][:4]: cell.border=Border(bottom=thin)
    summary.merge_cells("A17:H17")
    summary["A17"]="La hoja ‘Participantes y pagos’ contiene el detalle completo y permite filtrar los registros."
    summary["A17"].font=Font(italic=True,color="594A00")
    summary["A17"].fill=PatternFill("solid",fgColor="FFF7CC")
    summary["A17"].alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    for column in "ABCD": summary.column_dimensions[column].width=23
    for column in "EFGH": summary.column_dimensions[column].width=4
    summary.freeze_panes="A4"

    headers=["Código de solicitud","Nombre completo","Número de WhatsApp","Tipo de documento",
             "Número de documento","Cantidad","Total (S/)","Omar (S/)","Franci (S/)","Estado","Tickets confirmados/asignados",
             "Tickets originales/liberados","Fecha de registro","Fecha de revisión","Observación",
             "Reactivada desde vencida"]
    detail.append(headers)
    for row in rows:
        current_tickets=row.get("tickets",[]) or []
        historical=row.get("original_tickets",[]) or row.get("released_tickets",[]) or current_tickets
        approved_total=int(row.get("total",0)) if row.get("status")=="approved" else None
        amount_omar=row.get("payment_amount_omar")
        amount_franci=row.get("payment_amount_franci")
        # Compatibilidad con pagos aprobados antes de permitir montos divididos.
        if approved_total is not None and amount_omar is None and amount_franci is None:
            amount_omar=approved_total if row.get("payment_recipient")=="omar" else None
            amount_franci=approved_total if row.get("payment_recipient")=="franci" else None
        detail.append([
            safe_excel_text(row.get("code")),safe_excel_text(row.get("name")),safe_excel_text(row.get("phone")),
            safe_excel_text(row.get("document_type")),safe_excel_text(row.get("document_number")),
            int(row.get("quantity",0)),int(row.get("total",0)),
            int(amount_omar) if amount_omar not in {None,0} else None,
            int(amount_franci) if amount_franci not in {None,0} else None,
            status_labels.get(row.get("status"),row.get("status","")),
            ", ".join(str(number) for number in current_tickets),
            ", ".join(str(number) for number in historical),
            local_datetime(row.get("created_at")),local_datetime(row.get("reviewed_at")),
            safe_excel_text(row.get("review_note")),"Sí" if row.get("reactivated_from_expired") else "No",
        ])

    for cell in detail[1]:
        cell.font=Font(bold=True,color=white)
        cell.fill=PatternFill("solid",fgColor=dark)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        cell.border=Border(bottom=thin)
    for row in detail.iter_rows(min_row=2):
        status=row[9].value
        for cell in row:
            cell.alignment=Alignment(vertical="top",wrap_text=True)
            cell.border=Border(bottom=thin)
        if row[0].row%2==0:
            for cell in row: cell.fill=PatternFill("solid",fgColor="F8FAFC")
        status_colors={"Aprobada":("DCFCE7","166534"),"Pendiente":("FEF3C7","92400E"),
                       "Rechazada":("FEE2E2","991B1B"),"Vencida":("E5E7EB","374151")}
        if status in status_colors:
            fill_color,font_color=status_colors[status]
            row[9].fill=PatternFill("solid",fgColor=fill_color)
            row[9].font=Font(bold=True,color=font_color)
        for currency_cell in row[6:9]: currency_cell.number_format='"S/" #,##0'
        row[12].number_format='dd/mm/yyyy hh:mm'
        row[13].number_format='dd/mm/yyyy hh:mm'
    widths=[20,32,20,18,20,12,14,14,14,16,30,30,21,21,48,23]
    for index,width in enumerate(widths,1):
        detail.column_dimensions[chr(64+index)].width=width
    detail.freeze_panes="A2"
    detail.auto_filter.ref=f"A1:P{max(detail.max_row,1)}"

    output=BytesIO()
    workbook.save(output)
    output.seek(0)
    filename="chapa-tu-suerte-participantes-"+datetime.now(lima).strftime("%Y%m%d-%H%M")+".xlsx"
    return send_file(output,as_attachment=True,download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.get("/admin/proof/<filename>")
@admin_required
def proof(filename):
    # filename comes from our own generated code, not arbitrary user input.
    if Path(filename).name != filename: return "bad",400
    try:
        blob=storage_bucket().blob(filename)
        raw=blob.download_as_bytes()
        response=make_response(raw)
        response.headers["Content-Type"]=blob.content_type or mimetypes.guess_type(filename)[0] or "application/octet-stream"
        # Permite que el navegador reutilice el comprobante durante la sesión.
        response.headers["Cache-Control"]="private, max-age=300"
        return response
    except Exception:
        return "No se pudo cargar el comprobante",404

@app.post("/api/admin/orders/<oid>/approve")
@admin_required
@roles_allowed("owner","validator")
def approve(oid):
    ref=firestore_db().collection("orders").document(oid)
    snapshot=ref.get()
    if not snapshot.exists: return jsonify(error="No encontrado"),404
    order=snapshot.to_dict()
    if order["status"] not in {"pending","expired"}:
        return jsonify(error="La solicitud ya fue procesada"),400
    now=int(time.time())
    reviewed_at_text=time.strftime("%Y-%m-%d %H:%M:%S",time.gmtime(now))
    requested_note=(request.json or {}).get("note") or "Pago validado"
    data=request.json or {}
    try:
        amount_omar=int(data.get("amount_omar",0)); amount_franci=int(data.get("amount_franci",0))
    except (TypeError,ValueError):
        return jsonify(error="Los montos de Omar y Franci deben ser números enteros"),400
    if amount_omar<0 or amount_franci<0 or amount_omar+amount_franci!=int(order.get("total",0)):
        return jsonify(error=f"Los montos de Omar y Franci deben sumar S/{order.get('total',0)}"),400

    if order["status"]=="pending":
        batch=firestore_db().batch()
        batch.update(ref, {"status":"approved", "reviewed_at":now,
                           "reviewed_at_text":reviewed_at_text,
                           "review_note":requested_note,"payment_amount_omar":amount_omar,
                           "payment_amount_franci":amount_franci,"payment_recipient":None})
        for number in order.get("tickets",[]):
            batch.update(firestore_db().collection("tickets").document(str(number)), {"status":"confirmed"})
        batch.commit()
        audit("order_approved",{"order_id":oid,"order_code":order.get("code")})
        return jsonify(ok=True,message="Pago aprobado y tickets confirmados",tickets=order.get("tickets",[]))

    # Una reserva vencida ya liberó sus números. Recuperamos los disponibles
    # y reemplazamos únicamente los que hayan sido asignados a otra compra.
    original=[int(number) for number in order.get("released_tickets",[]) or order.get("tickets",[])]
    quantity=int(order.get("quantity",len(original)))
    tickets_collection=firestore_db().collection("tickets")
    for _ in range(12):
        replacement_pool=[]
        while len(replacement_pool)<max(quantity*3,20):
            number=secrets.randbelow(900000)+100000
            if number not in original and number not in replacement_pool:
                replacement_pool.append(number)

        transaction=firestore_db().transaction()
        @firestore.transactional
        def restore_expired(transaction):
            current=ref.get(transaction=transaction)
            if not current.exists or current.to_dict().get("status")!="expired":
                return None
            candidates=original+replacement_pool
            candidate_refs=[tickets_collection.document(str(number)) for number in candidates]
            candidate_snapshots=[ticket_ref.get(transaction=transaction) for ticket_ref in candidate_refs]
            available={number for number,ticket_snapshot in zip(candidates,candidate_snapshots)
                       if not ticket_snapshot.exists}
            selected=[number for number in original if number in available]
            selected.extend(number for number in replacement_pool
                            if number in available and len(selected)<quantity)
            if len(selected)<quantity:
                return False
            selected=selected[:quantity]
            replaced=[number for number in original if number not in selected]
            replacements=[number for number in selected if number not in original]
            history_note=requested_note
            if replacements:
                history_note+=(". Reserva vencida reactivada; números reemplazados: "
                               +", ".join(str(number) for number in replaced)
                               +" por "+", ".join(str(number) for number in replacements))
            else:
                history_note+=". Reserva vencida reactivada con sus números originales"
            transaction.update(ref,{"status":"approved", "tickets":selected,
                                    "reviewed_at":now, "reviewed_at_text":reviewed_at_text,
                                    "review_note":history_note,
                                    "reactivated_from_expired":True,
                                    "original_tickets":original,"payment_amount_omar":amount_omar,
                                    "payment_amount_franci":amount_franci,"payment_recipient":None})
            for number in selected:
                transaction.set(tickets_collection.document(str(number)),{
                    "number":number, "order_id":oid, "order_code":order.get("code"),
                    "status":"confirmed", "created_at":now,
                })
            return {"tickets":selected,"replaced":replaced,"replacements":replacements}

        result=restore_expired(transaction)
        if result is None:
            return jsonify(error="La solicitud ya fue procesada"),400
        if result:
            message="Pago vencido aprobado con los números originales"
            if result["replacements"]:
                message="Pago vencido aprobado; se asignaron números de reemplazo"
            audit("expired_order_approved",{"order_id":oid,"order_code":order.get("code")})
            return jsonify(ok=True,message=message,**result)
    return jsonify(error="No se pudieron reservar números disponibles. Intenta nuevamente"),409

@app.post("/api/admin/orders/<oid>/reject")
@admin_required
@roles_allowed("owner","validator")
def reject(oid):
    ref=firestore_db().collection("orders").document(oid)
    snapshot=ref.get()
    if not snapshot.exists: return jsonify(error="No encontrado"),404
    order=snapshot.to_dict()
    if order["status"]!="pending": return jsonify(error="La solicitud ya fue procesada"),400
    now=int(time.time())
    batch=firestore_db().batch()
    batch.update(ref, {"status":"rejected", "reviewed_at":now,
                       "reviewed_at_text":time.strftime("%Y-%m-%d %H:%M:%S",time.gmtime(now)),
                       "review_note":(request.json or {}).get("note") or "Pago rechazado",
                       "payment_recipient":None,
                       "payment_amount_omar":None,"payment_amount_franci":None,
                       "released_tickets":order.get("tickets",[]), "tickets":[]})
    for number in order.get("tickets",[]):
        batch.delete(firestore_db().collection("tickets").document(str(number)))
    batch.commit()
    audit("order_rejected",{"order_id":oid,"order_code":order.get("code")})
    return jsonify(ok=True)

@app.patch("/api/admin/orders/<oid>/recipient")
@admin_required
@roles_allowed("owner","validator")
def update_payment_recipient(oid):
    ref=firestore_db().collection("orders").document(oid)
    snapshot=ref.get()
    if not snapshot.exists: return jsonify(error="Solicitud no encontrada"),404
    order=snapshot.to_dict() or {}
    if order.get("status")!="approved":
        return jsonify(error="Solo se puede asignar la cuenta a pagos aprobados"),400
    data=request.json or {}
    try:
        amount_omar=int(data.get("amount_omar",0)); amount_franci=int(data.get("amount_franci",0))
    except (TypeError,ValueError):
        return jsonify(error="Los montos deben ser números enteros"),400
    if amount_omar<0 or amount_franci<0 or amount_omar+amount_franci!=int(order.get("total",0)):
        return jsonify(error=f"Omar y Franci deben sumar S/{order.get('total',0)}"),400
    ref.update({"payment_amount_omar":amount_omar,"payment_amount_franci":amount_franci,
                "payment_recipient":None,"recipient_updated_at":int(time.time()),
                "recipient_updated_by":g.admin_user})
    audit("payment_recipient_updated",{"order_id":oid,"order_code":order.get("code"),
                                        "amount_omar":amount_omar,"amount_franci":amount_franci})
    return jsonify(ok=True,amount_omar=amount_omar,amount_franci=amount_franci)

@app.post("/api/admin/orders/delete-selected")
@admin_required
@roles_allowed("owner")
def delete_selected_orders():
    data=request.json or {}
    raw_ids=data.get("ids") or []
    if not isinstance(raw_ids,list):
        return jsonify(error="Selección inválida"),400
    ids=list(dict.fromkeys(str(value).strip() for value in raw_ids))
    if not ids or len(ids)>50 or any(not re.fullmatch(r"\d{10,24}",oid) for oid in ids):
        return jsonify(error="Selecciona entre 1 y 50 compras válidas"),400
    if data.get("confirmation")!="ELIMINAR":
        return jsonify(error="Confirmación inválida"),400

    db=firestore_db()
    deleted=[]
    not_found=[]
    proof_warnings=[]
    for oid in ids:
        order_ref=db.collection("orders").document(oid)
        snapshot=order_ref.get()
        if not snapshot.exists:
            not_found.append(oid)
            continue
        order=snapshot.to_dict() or {}

        # Solo elimina tickets que todavía pertenecen a esta compra. Los
        # números liberados que otra persona recibió nunca se tocan.
        owned_tickets=list(
            db.collection("tickets").where("order_id","==",oid).limit(100).stream()
        )
        batch=db.batch()
        for ticket_snapshot in owned_tickets:
            batch.delete(ticket_snapshot.reference)
        batch.delete(order_ref)
        batch.commit()
        deleted.append(oid)

        proof_paths=order.get("proof_paths") or ([order.get("proof_path")] if order.get("proof_path") else [])
        for proof_path in dict.fromkeys(path for path in proof_paths if path):
            try:
                storage_bucket().blob(proof_path).delete()
            except Exception:
                # La compra ya fue retirada del sistema. GCS tiene eliminación
                # recuperable configurada y este aviso permite revisar el objeto.
                proof_warnings.append(oid)

    audit("orders_deleted",{"order_ids":deleted,"count":len(deleted)})
    return jsonify(ok=True,deleted=deleted,not_found=not_found,
                   proof_warnings=proof_warnings,count=len(deleted))

def valid_admin_password(password):
    return (len(password)>=14 and re.search(r"[a-z]",password) and
            re.search(r"[A-Z]",password) and re.search(r"\d",password) and
            re.search(r"[^A-Za-z0-9]",password))

def revoke_user_sessions(username):
    snapshots=list(firestore_db().collection("admin_sessions").where(
        "actor","==",username).limit(100).stream())
    if snapshots:
        batch=firestore_db().batch()
        for snapshot in snapshots:
            batch.delete(snapshot.reference)
            _ADMIN_SESSION_CACHE.pop(snapshot.id,None)
        batch.commit()
    return len(snapshots)

@app.get("/api/admin/users")
@admin_required
@roles_allowed("owner")
def list_admin_users():
    users=[{
        "username":ADMIN_USER,"role":"owner","role_label":"ADMINISTRADOR PRINCIPAL",
        "active":True,"source":"render","created_at":"Cuenta principal de Render",
        "last_login_at":"",
    }]
    for snapshot in firestore_db().collection("admin_users").limit(100).stream():
        row=snapshot.to_dict() or {}
        users.append({
            "username":row.get("username",snapshot.id),"role":row.get("role","viewer"),
            "role_label":{"validator":"VALIDADOR DE PAGOS","viewer":"SOLO CONSULTA"}.get(
                row.get("role"),"SOLO CONSULTA"),
            "active":bool(row.get("active",False)),"source":"firestore",
            "created_at":lima_datetime_text(row.get("created_at")),
            "last_login_at":lima_datetime_text(row.get("last_login_at")),
        })
    return jsonify(users=users,current_user=g.admin_user,current_role=g.admin_role)

@app.post("/api/admin/users")
@admin_required
@roles_allowed("owner")
def create_admin_user():
    import pyotp
    data=request.json or {}
    username=str(data.get("username","")).strip().lower()
    password=str(data.get("password", ""))
    role=str(data.get("role","")).strip()
    if not re.fullmatch(r"[a-z0-9._-]{3,32}",username):
        return jsonify(error="El usuario debe tener entre 3 y 32 caracteres: letras, números, punto, guion o guion bajo"),400
    if username==ADMIN_USER.lower():
        return jsonify(error="Ese usuario pertenece a la cuenta principal"),409
    if role not in {"validator","viewer"}:
        return jsonify(error="Selecciona un rol válido"),400
    if not valid_admin_password(password):
        return jsonify(error="La contraseña debe tener 14 caracteres, mayúscula, minúscula, número y símbolo"),400
    ref=firestore_db().collection("admin_users").document(username)
    if ref.get().exists:
        return jsonify(error="Ese nombre de usuario ya existe"),409
    secret=pyotp.random_base32()
    now=int(time.time())
    ref.set({
        "username":username,"password_hash":generate_password_hash(password),
        "totp_secret_encrypted":encrypt_totp(secret),"role":role,"active":True,
        "created_at":now,"created_by":g.admin_user,"last_login_at":None,
    })
    audit("admin_user_created",{"username":username,"role":role})
    uri=pyotp.TOTP(secret).provisioning_uri(name=username,issuer_name="Chapa Tu Suerte")
    return jsonify(ok=True,username=username,role=role,totp_secret=secret,provisioning_uri=uri),201

@app.patch("/api/admin/users/<username>")
@admin_required
@roles_allowed("owner")
def update_admin_user(username):
    username=str(username).strip().lower()
    if username==ADMIN_USER.lower():
        return jsonify(error="La cuenta principal se administra desde Render"),400
    ref=firestore_db().collection("admin_users").document(username)
    snapshot=ref.get()
    if not snapshot.exists: return jsonify(error="Usuario no encontrado"),404
    data=request.json or {}; updates={}
    if "active" in data: updates["active"]=bool(data.get("active"))
    if "role" in data:
        if data.get("role") not in {"validator","viewer"}:
            return jsonify(error="Rol inválido"),400
        updates["role"]=data["role"]
    if not updates: return jsonify(error="No hay cambios válidos"),400
    updates["updated_at"]=int(time.time()); updates["updated_by"]=g.admin_user
    ref.update(updates)
    closed=revoke_user_sessions(username)
    audit("admin_user_updated",{"username":username,"changes":list(updates.keys()),"sessions_closed":closed})
    return jsonify(ok=True,sessions_closed=closed)

@app.post("/api/admin/users/<username>/reset-security")
@admin_required
@roles_allowed("owner")
def reset_admin_security(username):
    import pyotp
    username=str(username).strip().lower()
    if username==ADMIN_USER.lower():
        return jsonify(error="La cuenta principal se administra desde Render"),400
    ref=firestore_db().collection("admin_users").document(username)
    if not ref.get().exists: return jsonify(error="Usuario no encontrado"),404
    password=str((request.json or {}).get("password", ""))
    if not valid_admin_password(password):
        return jsonify(error="La contraseña debe tener 14 caracteres, mayúscula, minúscula, número y símbolo"),400
    secret=pyotp.random_base32()
    ref.update({"password_hash":generate_password_hash(password),
                "totp_secret_encrypted":encrypt_totp(secret),"active":True,
                "security_reset_at":int(time.time()),"security_reset_by":g.admin_user})
    closed=revoke_user_sessions(username)
    audit("admin_user_security_reset",{"username":username,"sessions_closed":closed})
    uri=pyotp.TOTP(secret).provisioning_uri(name=username,issuer_name="Chapa Tu Suerte")
    return jsonify(ok=True,username=username,totp_secret=secret,provisioning_uri=uri)

@app.get("/api/admin/stats")
@admin_required
def stats():
    rows=[snapshot.to_dict() for snapshot in firestore_db().collection("orders").stream()]
    out={status:sum(1 for row in rows if row.get("status")==status)
         for status in ("pending","approved","rejected","expired")}
    approved=[row for row in rows if row.get("status")=="approved"]
    pending=[row for row in rows if row.get("status")=="pending"]
    out["confirmed_tickets"]=sum(int(row.get("quantity",0)) for row in approved)
    out["reserved_tickets"]=sum(int(row.get("quantity",0)) for row in pending)
    out["sales_approved"]=sum(int(row.get("total",0)) for row in approved)
    return jsonify(out)

@app.get("/admin")
def admin_page():
    return send_from_directory(BASE,"admin.html")

if __name__=="__main__":
    app.run(host="0.0.0.0",port=int(os.environ.get("PORT",5000)),debug=False)
